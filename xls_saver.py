from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelSaver:
    def __init__(self,file_name):
        self.file_name = file_name
        self.wb = Workbook()
        self.sheet1 = self.wb.active
        self.sheet1.title = 'etherscan.io'
        self.sheet1.append(['TxHash', 'From', 'To', 'Page', 'TimeStamp'])
        self.wb.save(file_name)
        self.hash_set = set()

    def save_to_file(self, hashs_queue):
        wb = load_workbook(self.file_name)
        sheet1_append = wb.active
        max_row = self.sheet1.max_row
        # write every datetime for each page

        self.wb.save(filename=self.file_name)
        is_first = True
        datas = hashs_queue.get()
        print(datas)

        hashs = datas[0]
        page = datas[1]
        times = datas[2]

        # the pages displayed Txhash - from order in hash-tag text-truncate
        # make tuple collection from hashs idx modular
        for tx_hash, tx_from, tx_to in zip(hashs[0::3], hashs[1::3], hashs[2::3]):

            # for avoid duplicate
            if tx_hash.text in self.hash_set:
                pass
            else:
                sheet1_append.append([tx_hash.text, tx_from.text,tx_to.text, page, times])
                self.hash_set.add(tx_hash.text)
                is_first = False

        wb.save(self.file_name)

